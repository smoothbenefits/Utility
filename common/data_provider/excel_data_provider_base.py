from openpyxl import load_workbook


class ExcelDataProviderBase(object):
    def __init__(self, excel_file):
        self._excel_file = excel_file

    ''' This should return an instance of ModelParserBase
    '''
    def _get_row_parser(self):
        raise NotImplementedError("Please Implement this method in the sub-class")

    def _get_header_row(self, work_sheet):
        header_row_num = 1
        header_row = None
        for row in work_sheet.rows:
            header_row = row
            break
        return header_row, header_row_num

    def __is_empty_row(self, row):
        for cell in row:
            if cell.value:
                return False
        return True

    def __get_work_sheet(self):
        excel_file = self._excel_file
        if excel_file:
            work_book = load_workbook(excel_file, read_only=True, data_only=True, keep_vba=True)
            work_sheet = work_book.active
            return work_sheet
        else:
            return None

    def __parse_all_rows(self):
        parsed_models = []
        work_sheet = self.__get_work_sheet()
        header_row, header_row_num = self._get_header_row(work_sheet)
        row_parser = self._get_row_parser()
        row_parser.initialize(header_row)

        for row in work_sheet.iter_rows(row_offset=header_row_num):
            if (not row or self.__is_empty_row(row)):
                continue
            parsed_model = row_parser.parse_data_row(row)
            parsed_models.append(parsed_model)

        return parsed_models

    def get_model(self):
        return self.__parse_all_rows()
