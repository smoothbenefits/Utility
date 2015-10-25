from openpyxl import load_workbook

from parser.fairview.address_parser import AddressParser


class UserDataProvider(object):
    def __init__(self, excel_file):
        self.work_book = load_workbook(excel_file, read_only=True, data_only=True, keep_vba=True)
        self.work_sheet = self.work_book.active

    def process(self):
        # initialize parsers
        self.parsers = []
        self.parsers.append(AddressParser())

        header_row, self.header_row_num = self._find_header_row(self.work_sheet)
        for parser in self.parsers:
            parser.initialize_column_mapping(header_row)

        # start processing data
        self.models = []
        self._process_data()

        # Serialize. This should be break out of the class.
        # Just put it here for quicker playing for now
        self._serialize_data()

    def _find_header_row(self, work_sheet):
        header_row = None
        for row in work_sheet.rows:
            header_row = row
            for cell in row:
                if cell.value =='First Name' or cell.value == 'Last Name':
                    return header_row, cell.row
        return None, None

    def _process_data(self):
        for row in self.work_sheet.iter_rows(row_offset=self.header_row_num):
            if not row:
                continue
            model = self.parsers[0].parse_data_row(row)
            self.models.append(model)

    def _serialize_data(self):
        for model in self.models:
            print "{}, {}".format(model.street_1, model.street_2)
