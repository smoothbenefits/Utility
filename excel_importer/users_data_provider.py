from openpyxl import load_workbook
from parser.row_parser import RowParser
from model.excel_type import ExcelType

class UsersDataProvider(object):
    def __init__(self, company_users):
        self.company_users = company_users

    def process(self, excel_file):
        work_book = load_workbook(excel_file, read_only=True, data_only=True, keep_vba=True)
        work_sheet = work_book.active
        # initialize parsers
        header_row, header_row_num = self._find_header_row(work_sheet)
        row_parser = RowParser()
        row_parser.initialize(header_row)

        for row in work_sheet.iter_rows(row_offset=header_row_num):
            if not row:
                continue
            row_parsed = row_parser.parse_data(row)
            self.company_users.merge_with_excel_data(row_parsed, self._get_excel_type(excel_file))

    def _find_header_row(self, work_sheet):
        header_row = None
        for row in work_sheet.rows:
            header_row = row
            for cell in row:
                if cell.value =='First Name' or cell.value == 'Last Name':
                    return header_row, cell.row
        return None, None

    def _get_excel_type(self, file_name):
        if ExcelType.HCHP in file_name:
            return ExcelType.HCHP
        else:
            return ExcelType.ENROLLMENT



