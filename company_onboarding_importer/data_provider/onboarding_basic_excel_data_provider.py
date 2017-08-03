from openpyxl import load_workbook
from ..excel_parser.basic.row_parser import RowParser

class OnboardingBasicExcelDataProvider(object):
    def provide(self, excel_file, sheet_name, company_users):
        work_book = load_workbook(excel_file, read_only=True, data_only=True, keep_vba=True)
        work_sheet = work_book.get_sheet_by_name(sheet_name)
        # initialize parsers
        header_row, header_row_num = self._find_header_row(work_sheet)
        row_parser = RowParser()
        row_parser.initialize(header_row)

        for row in work_sheet.iter_rows(row_offset=header_row_num):
            if not row:
                continue
            row_parsed = row_parser.parse_data(row)
            company_users.merge_with_file_data(row_parsed)

    def _find_header_row(self, work_sheet):
        header_row = None
        for row in work_sheet.rows:
            header_row = row
            for cell in row:
                if cell.value =='First Name' or cell.value == 'Email':
                    return header_row, cell.row
        return None, None
